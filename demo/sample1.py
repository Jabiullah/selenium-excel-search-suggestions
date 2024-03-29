import datetime
import time
from selenium import webdriver
from selenium.webdriver import Keys, ActionChains
from openpyxl import load_workbook

current_day_index = datetime.datetime.now().weekday()
print("Today is:", current_day_index)

days_to_sheet = {
    0: "Monday",
    1: "Tuesday",
    2: "Wednesday",
    3: "Thursday",
    4: "Friday",
    5: "Saturday",
    6: "Sunday"
}

sheet_name = days_to_sheet[current_day_index]

workbook = load_workbook("C:\\Users\\ASUS\\PycharmProjects\\seleniumTest\\Excel.xlsx")

sheet = workbook[sheet_name]
print(sheet)
search_keys = []
for row in range(3, 13):  # Assuming C3 to C12
    search_key = sheet.cell(row=row, column=3).value
    if search_key:
        search_keys.append(search_key)

workbook.close()

driver = webdriver.Chrome()
driver.maximize_window()
driver.get("https://www.google.com/")

for i,search_key in enumerate(search_keys):
    search_box = driver.find_element("name", "q")
    search_box.clear()
    search_box.send_keys(search_key)
    ActionChains(driver).move_to_element(search_box).perform()
    time.sleep(3)
    suggestion_elements = driver.find_elements("xpath","//ul[@role='listbox']//li")
    suggestions = []
    for element in suggestion_elements:
        suggestions.append(element.text)
    longest_word = max(suggestions, key=len)
    smallest_word = min(suggestions, key=len)
    print("Longest word for", search_key, ":", longest_word)
    print("Smallest word for", search_key, ":", smallest_word)
    sheet.cell(row=3 + i, column=4).value = longest_word
    sheet.cell(row=3 + i, column=5).value = smallest_word
    search_box.send_keys(Keys.ENTER)
    time.sleep(3)
workbook.save("C:\\Users\\ASUS\\PycharmProjects\\seleniumTest\\Excel.xlsx")
driver.quit()